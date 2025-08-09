import os, sys, json, datetime as dt
from pathlib import Path
from typing import Dict
import pandas as pd
import requests, gspread
from gspread.exceptions import WorksheetNotFound
from google.oauth2.service_account import Credentials
from dotenv import load_dotenv
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parents[1]
load_dotenv(dotenv_path=BASE_DIR / ".env", override=True)

def _abs(p: str, default: str = "") -> str:
    p = p or default
    return str((Path(p) if os.path.isabs(p) else (BASE_DIR / p)).resolve())

SHEET_ID = os.getenv("SHEET_ID", "").strip()
WS_NAME = (os.getenv("SHEET_WORKSHEET_NAME", "") or "").strip()
CREDS_PATH = _abs(os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON", "./credentials.json"), "./credentials.json")
OUTPUT_PATH = _abs(os.getenv("REPORT_PATH", "docs/daily_report.xlsx"), "docs/daily_report.xlsx")
SLACK_WEBHOOK= os.getenv("SLACK_WEBHOOK_URL", "").strip()
TODAY = dt.date.today()


COLORS = {
    "header": "EDF2FF", "sep": "F6F7FB", "text": "111827",
    "pill_delivered": "34C759", "pill_delayed": "FF6B6B",
    "pill_transit": "FFD166", "pill_received": "7AA8FF", "pill_shipped": "FFB020",
}
THIN = Side(style="thin", color="D9DCE3")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left",   vertical="center")
BOLD = Font(bold=True, color=COLORS["text"])
WHITE= Font(bold=True, color="FFFFFF")
STATUS_COLOR = {
    "delivered": COLORS["pill_delivered"],
    "delayed": COLORS["pill_delayed"],
    "in transit":COLORS["pill_transit"],
    "received":COLORS["pill_received"],
    "shipped": COLORS["pill_shipped"],
}

def style_header(cells):
    for c in cells:
        c.fill = PatternFill("solid", fgColor=COLORS["header"])
        c.font, c.alignment, c.border = BOLD, CENTER, BORDER

def style_cell(cell, align=CENTER, fill=None, font=None, border=BORDER):
    if fill: cell.fill = fill
    if font: cell.font = font
    cell.alignment = align
    cell.border = border

def auto_width(ws, cols):
    for col in cols:
        letter = get_column_letter(col)
        width = max(len(str(ws.cell(r, col).value or "")) for r in range(1, ws.max_row+1)) + 3
        ws.column_dimensions[letter].width = min(42, width)


def _gc():
    if not (SHEET_ID and os.path.exists(CREDS_PATH)):
        raise RuntimeError("Missing SHEET_ID or credentials json (GOOGLE_SERVICE_ACCOUNT_JSON).")
    creds = Credentials.from_service_account_file(CREDS_PATH,
             scopes=["https://www.googleapis.com/auth/spreadsheets.readonly"])
    return gspread.authorize(creds)

def load_df() -> pd.DataFrame:
    sh = _gc().open_by_key(SHEET_ID)
    try: ws = sh.worksheet(WS_NAME) if WS_NAME else sh.sheet1
    except WorksheetNotFound: ws = sh.sheet1

    df = pd.DataFrame(ws.get_all_records())
    df.columns = [c.strip() for c in df.columns]
    norm = {c.lower(): c for c in df.columns}
    req = ["orderid","customer","status","eta","email"]
    missing = [k for k in req if k not in norm];  assert not missing, f"Missing columns: {missing}"

    for col in ["orderid","customer","status"]: df[norm[col]] = df[norm[col]].astype(str).str.strip()
    df[norm["email"]] = df[norm["email"]].astype(str).str.lower().str.strip()
    df[norm["eta"]]   = pd.to_datetime(df[norm["eta"]], errors="coerce").dt.date

    pretty = {"orderid":"OrderID","customer":"Customer","status":"Status","eta":"ETA","email":"Email"}
    return df.rename(columns={norm[k]: v for k,v in pretty.items()})[list(pretty.values())]

def build_tabs(df: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    total = len(df)
    by_status = (df["Status"].value_counts()
                 .rename_axis("Status").reset_index(name="Count")
                 .sort_values("Status").reset_index(drop=True))
    delayed_mask = df["Status"].str.contains("delayed", case=False, na=False)
    pct_delayed = round(100 * delayed_mask.mean(), 1)
    sla_miss_cnt = int(((df["ETA"].notna()) & (df["ETA"] < TODAY) & (~df["Status"].str.contains("delivered", case=False))).sum())
    kpi = pd.DataFrame({"Metric": ["Total orders","Delayed %","SLA misses"],
                        "Value":  [total, f"{pct_delayed}%", sla_miss_cnt]})
    return {"KPI": kpi, "ByStatus": by_status}


def save_excel(tabs: Dict[str, pd.DataFrame], path: str):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as xl:
        wb = xl.book
        ws = wb.create_sheet("Daily") if "Daily" not in wb.sheetnames else wb["Daily"]
        ws.delete_rows(1, ws.max_row)

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
        title = ws.cell(1,1, f"Daily Logistics Report - {TODAY.isoformat()}")
        title.font, title.alignment = Font(bold=True, size=14, color=COLORS["text"]), LEFT

        ws.cell(3,1,"Metric"); ws.cell(3,2,"Value"); style_header(ws[3][0:2])
        r = 4
        for m, v in tabs["KPI"].itertuples(index=False):
            style_cell(ws.cell(r,1,m), LEFT)
            style_cell(ws.cell(r,2,v))
            r += 1

        r += 1
        ws.cell(r,1,"Status"); ws.cell(r,2,"Count"); style_header(ws[r][0:2]); r += 1
        for status, cnt in tabs["ByStatus"].itertuples(index=False):
            s = status.strip(); fill = PatternFill("solid", fgColor=STATUS_COLOR.get(s.lower(), COLORS["sep"]))
            style_cell(ws.cell(r,1,s), CENTER, fill, WHITE)
            style_cell(ws.cell(r,2,int(cnt)))
            r += 1

        for i in range(3, r): ws.row_dimensions[i].height = 22
        auto_width(ws, [1,2])
        if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"])
        xl._save()

def slack_notify(tabs: Dict[str, pd.DataFrame]):
    if not SLACK_WEBHOOK: return
    kv = {row["Metric"]: str(row["Value"]) for _, row in tabs["KPI"].iterrows()}
    bys = tabs["ByStatus"]
    lines = [
        f"*Daily Logistics Report* - {TODAY.isoformat()}",
        f"• Total orders: *{kv.get('Total orders','?')}*",
        f"• Delayed %: *{kv.get('Delayed %','?')}*",
        f"• SLA misses: *{kv.get('SLA misses','?')}*",
        "• By status: " + ", ".join(f"{r.Status}: {r.Count}" for _, r in bys.iterrows()),
    ]
    try:
        r = requests.post(SLACK_WEBHOOK, data=json.dumps({"text":"\n".join(lines)}),
                          headers={"Content-Type":"application/json"}, timeout=10)
        r.raise_for_status()
    except Exception as e:
        print(f"[slack] failed: {e}", file=sys.stderr)

def main():
    df = load_df()
    tabs = build_tabs(df)
    save_excel(tabs, OUTPUT_PATH)
    slack_notify(tabs)

if __name__ == "__main__":
    main()
